"""
Excel utilities for Sherpa Check-In.

Manages an Excel workbook with two sheets:
- Intake: for all client intakes (kiosk, drop-off, email, etc.)
- MailLog: for tracking outbound mail/documents to clients
"""

import os
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

logger = logging.getLogger('client_checkin')

# Workbook path from environment or default
EXCEL_PATH = os.environ.get('INTAKE_EXCEL_PATH', r'Y:\Clients\Client_Log.xlsx')

# Sheet headers
INTAKE_HEADERS = [
    'IntakeID',
    'DateTimeReceived',
    'ClientName',
    'ClientEmail',
    'ClientPhone',
    'Professional',
    'IntakeType',
    'DueDate',
    'Status',
    'CompletedDate',
    'Notes',
]

MAILLOG_HEADERS = [
    'MailID',
    'DateSent',
    'ClientName',
    'Professional',
    'ItemType',
    'Method',
    'TrackingNumber',
    'SentBy',
    'Notes',
]


def _ensure_workbook():
    """
    Ensure the workbook exists with both sheets and proper headers.
    Creates the file if it doesn't exist, or adds missing sheets/headers.
    Returns the workbook object (caller should save if changes were made).
    """
    if os.path.exists(EXCEL_PATH):
        try:
            wb = load_workbook(EXCEL_PATH)
        except InvalidFileException:
            logger.error('Invalid Excel file at %s, creating new workbook', EXCEL_PATH)
            wb = Workbook()
    else:
        wb = Workbook()

    # Ensure Intake sheet exists with headers
    if 'Intake' not in wb.sheetnames:
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) == 1:
            # Rename default sheet
            ws = wb['Sheet']
            ws.title = 'Intake'
        else:
            ws = wb.create_sheet('Intake')
        ws.append(INTAKE_HEADERS)
        logger.info('Created Intake sheet in workbook')
    else:
        ws = wb['Intake']
        # Check if headers exist (row 1)
        if ws.max_row == 0 or ws.cell(1, 1).value != 'IntakeID':
            ws.insert_rows(1)
            for col, header in enumerate(INTAKE_HEADERS, start=1):
                ws.cell(1, col, header)

    # Ensure MailLog sheet exists with headers
    if 'MailLog' not in wb.sheetnames:
        ws = wb.create_sheet('MailLog')
        ws.append(MAILLOG_HEADERS)
        logger.info('Created MailLog sheet in workbook')
    else:
        ws = wb['MailLog']
        # Check if headers exist (row 1)
        if ws.max_row == 0 or ws.cell(1, 1).value != 'MailID':
            ws.insert_rows(1)
            for col, header in enumerate(MAILLOG_HEADERS, start=1):
                ws.cell(1, col, header)

    # Remove default 'Sheet' if both Intake and MailLog exist
    if 'Sheet' in wb.sheetnames and 'Intake' in wb.sheetnames and 'MailLog' in wb.sheetnames:
        del wb['Sheet']

    return wb


def append_intake_row(
    intake_id: int,
    client_name: str,
    professional: str,
    intake_type: str,
    client_email: str = "",
    client_phone: str = "",
    due_date=None,  # date or datetime, can be None
    status: str = "Not started",
    notes: str = "",
) -> None:
    """
    Append a row to the Intake sheet.

    Args:
        intake_id: The check-in/intake ID from the database
        client_name: Client's name
        professional: Name of the professional assigned
        intake_type: Type of intake (Appointment, Drop-off, Email, Portal Upload, Mail-in)
        client_email: Client's email (optional)
        client_phone: Client's phone (optional)
        due_date: Due date for the intake (optional, date or datetime)
        status: Current status (default "Not started")
        notes: Additional notes (optional)
    """
    wb = _ensure_workbook()
    ws = wb['Intake']

    # Format due_date if provided
    due_date_value = None
    if due_date is not None:
        if hasattr(due_date, 'strftime'):
            due_date_value = due_date
        else:
            due_date_value = str(due_date)

    row = [
        intake_id,
        datetime.now(),
        client_name,
        client_email or "",
        client_phone or "",
        professional,
        intake_type,
        due_date_value,
        status,
        None,  # CompletedDate - empty initially
        notes or "",
    ]

    ws.append(row)
    wb.save(EXCEL_PATH)
    logger.info('Appended intake row %d to Excel', intake_id)


def append_mail_row(
    mail_id: int,
    client_name: str,
    professional: str,
    item_type: str,
    method: str,
    tracking_number: str = "",
    sent_by: str = "",
    notes: str = "",
) -> None:
    """
    Append a row to the MailLog sheet.

    Args:
        mail_id: The mail record ID from the database
        client_name: Client's name
        professional: Name of the professional
        item_type: Type of item sent (Original Return, Amended Return, etc.)
        method: Delivery method (USPS, FedEx, UPS, etc.)
        tracking_number: Tracking number if applicable (optional)
        sent_by: Initials of person who sent (optional)
        notes: Additional notes (optional)
    """
    wb = _ensure_workbook()
    ws = wb['MailLog']

    row = [
        mail_id,
        datetime.now().date(),
        client_name,
        professional,
        item_type,
        method,
        tracking_number or "",
        sent_by or "",
        notes or "",
    ]

    ws.append(row)
    wb.save(EXCEL_PATH)
    logger.info('Appended mail row %d to Excel', mail_id)
